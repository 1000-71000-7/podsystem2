import io
import csv
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import Response, send_file, request, jsonify, g
from functools import wraps
import json


class ReportExporter:
    """Класс для экспорта отчётов в различные форматы"""

    @staticmethod
    def export_to_csv(data, filename):
        """Экспорт в CSV формат"""
        try:
            output = io.StringIO()

            if not data:
                output.write("Нет данных")
                output.seek(0)
                return Response(
                    output.getvalue(),
                    mimetype='text/csv; charset=utf-8-sig',
                    headers={'Content-Disposition': f'attachment; filename={filename}.csv'}
                )

            # Определяем заголовки
            if isinstance(data, list) and len(data) > 0:
                headers = list(data[0].keys())
            elif isinstance(data, dict):
                headers = list(data.keys())
                data = [data]
            else:
                output.write(str(data))
                output.seek(0)
                return Response(
                    output.getvalue(),
                    mimetype='text/csv; charset=utf-8-sig',
                    headers={'Content-Disposition': f'attachment; filename={filename}.csv'}
                )

            writer = csv.DictWriter(output, fieldnames=headers, delimiter=';')
            writer.writeheader()
            writer.writerows(data)

            output.seek(0)

            return Response(
                output.getvalue(),
                mimetype='text/csv; charset=utf-8-sig',
                headers={'Content-Disposition': f'attachment; filename={filename}.csv'}
            )
        except Exception as e:
            print(f"CSV export error: {e}")
            return jsonify({'error': str(e)}), 500

    @staticmethod
    def export_to_excel(data, filename):
        """Экспорт в Excel формат"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт"

            if not data:
                ws['A1'] = "Нет данных"
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'}
                )

            # Определяем заголовки
            if isinstance(data, list) and len(data) > 0:
                headers = list(data[0].keys())
            elif isinstance(data, dict):
                headers = list(data.keys())
                data = [data]
            else:
                ws['A1'] = str(data)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'}
                )

            # Стили для заголовков
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Записываем данные
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    ws.cell(row=row, column=col, value=value)

            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем в BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'}
            )
        except Exception as e:
            print(f"Excel export error: {e}")
            return jsonify({'error': str(e)}), 500

    @staticmethod
    def export_to_pdf(data, filename, title="Отчёт по метрикам"):
        """Экспорт в PDF формат"""
        try:
            output = io.BytesIO()

            # Создаём PDF документ
            doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)

            # Стили
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1a73e8'),
                spaceAfter=30
            )

            # Содержимое документа
            story = []

            # Заголовок
            story.append(Paragraph(title, title_style))
            story.append(Paragraph(f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
            story.append(Spacer(1, 20))

            if not data:
                story.append(Paragraph("Нет данных для отображения", styles['Normal']))
            else:
                # Подготовка данных для таблицы
                if isinstance(data, list) and len(data) > 0:
                    headers = list(data[0].keys())
                    table_data = [headers]
                    for item in data[:100]:  # Ограничиваем 100 строками для PDF
                        row = []
                        for header in headers:
                            value = item.get(header, '')
                            if isinstance(value, (dict, list)):
                                value = json.dumps(value, ensure_ascii=False)[:100]
                            elif isinstance(value, (int, float)):
                                value = str(value)
                            row.append(str(value)[:50])
                        table_data.append(row)
                elif isinstance(data, dict):
                    headers = list(data.keys())
                    table_data = [['Параметр', 'Значение']]
                    for key, value in data.items():
                        if isinstance(value, (dict, list)):
                            value = json.dumps(value, ensure_ascii=False)
                        table_data.append([str(key), str(value)])
                else:
                    story.append(Paragraph(str(data), styles['Normal']))
                    doc.build(story)
                    output.seek(0)
                    return send_file(
                        output,
                        download_name=f'{filename}.pdf',
                        as_attachment=True,
                        mimetype='application/pdf'
                    )

                # Создаём таблицу
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))

                story.append(table)

            # Строим документ
            doc.build(story)
            output.seek(0)

            return send_file(
                output,
                download_name=f'{filename}.pdf',
                as_attachment=True,
                mimetype='application/pdf'
            )
        except Exception as e:
            print(f"PDF export error: {e}")
            return jsonify({'error': str(e)}), 500


def export_format():
    """Декоратор для поддержки форматов экспорта"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            format_type = request.args.get('format', 'json')

            result = f(*args, **kwargs)

            # Если результат уже является Response, возвращаем его
            if isinstance(result, Response):
                return result

            # Экспортируем в нужном формате
            if format_type == 'csv':
                return ReportExporter.export_to_csv(result, 'report')
            elif format_type == 'xlsx':
                return ReportExporter.export_to_excel(result, 'report')
            elif format_type == 'pdf':
                return ReportExporter.export_to_pdf(result, 'report', 'Отчёт по метрикам')
            else:
                return jsonify(result)

        return decorated_function

    return decorator